import csv
import io
import json
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone

class DataExporter:
    """Export data to various formats"""
    
    @classmethod
    def export_to_csv(cls, queryset, fields, filename=None):
        """Export queryset to CSV"""
        if not filename:
            filename = f"export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow([fields[f] if isinstance(fields, dict) else f for f in fields])
        
        # Write data
        for obj in queryset:
            row = []
            for field in fields:
                if isinstance(fields, dict):
                    field_name = field
                else:
                    field_name = field
                
                value = cls.get_nested_value(obj, field_name)
                row.append(str(value) if value is not None else '')
            
            writer.writerow(row)
        
        return response
    
    @classmethod
    def export_to_json(cls, data, filename=None):
        """Export data to JSON"""
        if not filename:
            filename = f"export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        json.dump(data, response, indent=2, default=str, ensure_ascii=False)
        
        return response
    
    @classmethod
    def export_to_excel(cls, data, filename=None):
        """Export data to Excel"""
        import openpyxl
        
        if not filename:
            filename = f"export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if isinstance(data, list) and data:
            # Write headers
            for col, key in enumerate(data[0].keys(), 1):
                ws.cell(row=1, column=col, value=key)
            
            # Write data
            for row, item in enumerate(data, 2):
                for col, key in enumerate(item.keys(), 1):
                    ws.cell(row=row, column=col, value=item[key])
        elif isinstance(data, dict):
            # Write single row
            for col, (key, value) in enumerate(data.items(), 1):
                ws.cell(row=1, column=col, value=key)
                ws.cell(row=2, column=col, value=value)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    @classmethod
    def export_to_pdf(cls, html_content, filename=None):
        """Export HTML to PDF"""
        from weasyprint import HTML
        
        if not filename:
            filename = f"export_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        html = HTML(string=html_content)
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @classmethod
    def get_nested_value(cls, obj, field_path):
        """Get nested value from object using dot notation"""
        parts = field_path.split('.')
        value = obj
        
        for part in parts:
            if hasattr(value, part):
                value = getattr(value, part)
            elif isinstance(value, dict) and part in value:
                value = value[part]
            else:
                return None
            
            if callable(value):
                value = value()
        
        return value

class ReportExporter:
    """Export reports in different formats"""
    
    @classmethod
    def export_report(cls, report_data, format='csv'):
        """Export report in specified format"""
        if format == 'csv':
            return cls.to_csv(report_data)
        elif format == 'json':
            return cls.to_json(report_data)
        elif format == 'excel':
            return cls.to_excel(report_data)
        elif format == 'pdf':
            return cls.to_pdf(report_data)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    @classmethod
    def to_csv(cls, report_data):
        """Convert report to CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write summary
        if 'summary' in report_data:
            writer.writerow(['SUMMARY'])
            for key, value in report_data['summary'].items():
                writer.writerow([key, value])
            writer.writerow([])
        
        # Write data tables
        for key, value in report_data.items():
            if key != 'summary' and isinstance(value, list) and value:
                writer.writerow([key.upper()])
                if value and isinstance(value[0], dict):
                    writer.writerow(value[0].keys())
                    for item in value:
                        writer.writerow(item.values())
                writer.writerow([])
        
        return output.getvalue()
    
    @classmethod
    def to_json(cls, report_data):
        """Convert report to JSON"""
        return json.dumps(report_data, indent=2, default=str)
    
    @classmethod
    def to_excel(cls, report_data):
        """Convert report to Excel"""
        import openpyxl
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        
        row = 1
        if 'summary' in report_data:
            for key, value in report_data['summary'].items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                row += 1
        
        # Data sheets
        for key, value in report_data.items():
            if key != 'summary' and isinstance(value, list) and value:
                ws = wb.create_sheet(title=key[:30])
                
                if value and isinstance(value[0], dict):
                    # Headers
                    for col, header in enumerate(value[0].keys(), 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # Data
                    for row, item in enumerate(value, 2):
                        for col, header in enumerate(item.keys(), 1):
                            ws.cell(row=row, column=col, value=item[header])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @classmethod
    def to_pdf(cls, report_data):
        """Convert report to PDF"""
        from weasyprint import HTML
        
        # Generate HTML
        html = cls.generate_html(report_data)
        
        html_obj = HTML(string=html)
        pdf = html_obj.write_pdf()
        
        return pdf
    
    @classmethod
    def generate_html(cls, report_data):
        """Generate HTML from report data"""
        html = """
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>Analytics Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                h1 { color: #007bff; }
                h2 { color: #333; margin-top: 30px; }
                table { border-collapse: collapse; width: 100%; margin: 10px 0; }
                th { background: #007bff; color: white; padding: 10px; text-align: left; }
                td { padding: 8px; border-bottom: 1px solid #ddd; }
                .summary { background: #f8f9fa; padding: 15px; border-radius: 5px; }
                .metric { display: inline-block; margin: 10px; padding: 15px; background: white; 
                         border: 1px solid #ddd; border-radius: 5px; min-width: 150px; }
                .metric-value { font-size: 24px; font-weight: bold; color: #007bff; }
                .metric-label { color: #666; font-size: 12px; }
            </style>
        </head>
        <body>
            <h1>Analytics Report</h1>
            <p>Generated: """ + timezone.now().strftime('%Y-%m-%d %H:%M') + """</p>
        """
        
        # Summary
        if 'summary' in report_data:
            html += "<h2>Summary</h2>"
            html += '<div class="summary">'
            for key, value in report_data['summary'].items():
                html += f"""
                <div class="metric">
                    <div class="metric-value">{value}</div>
                    <div class="metric-label">{key}</div>
                </div>
                """
            html += '</div>'
        
        # Data tables
        for key, value in report_data.items():
            if key != 'summary' and isinstance(value, list) and value:
                html += f"<h2>{key.title()}</h2>"
                html += '<table>'
                
                if value and isinstance(value[0], dict):
                    # Headers
                    html += '<thead><tr>'
                    for header in value[0].keys():
                        html += f'<th>{header}</th>'
                    html += '</tr></thead>'
                    
                    # Data
                    html += '<tbody>'
                    for item in value:
                        html += '<tr>'
                        for val in item.values():
                            html += f'<td>{val}</td>'
                        html += '</tr>'
                    html += '</tbody>'
                
                html += '</table>'
        
        html += """
        </body>
        </html>
        """
        
        return html