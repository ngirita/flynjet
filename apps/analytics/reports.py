from django.db.models import Sum, Count, Avg
from django.utils import timezone
from .models import Report, DailyMetric, AnalyticsEvent
from apps.bookings.models import Booking
from apps.payments.models import Payment
from apps.accounts.models import User
import logging

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate various reports"""
    
    @classmethod
    def generate_revenue_report(cls, start_date, end_date):
        """Generate revenue report for date range"""
        payments = Payment.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            status='completed'
        )
        
        daily = DailyMetric.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date')
        
        return {
            'period': f"{start_date} to {end_date}",
            'summary': {
                'total_revenue': float(payments.aggregate(Sum('amount_usd'))['amount_usd__sum'] or 0),
                'total_payments': payments.count(),
                'avg_payment': float(payments.aggregate(Avg('amount_usd'))['amount_usd__avg'] or 0),
            },
            'by_method': list(payments.values('payment_method').annotate(
                total=Sum('amount_usd'),
                count=Count('id')
            )),
            'daily': [{
                'date': m.date,
                'revenue': float(m.revenue),
                'bookings': m.total_bookings
            } for m in daily],
            'chart_data': {
                'labels': [m.date.strftime('%Y-%m-%d') for m in daily],
                'datasets': [{
                    'label': 'Daily Revenue',
                    'data': [float(m.revenue) for m in daily]
                }]
            }
        }
    
    @classmethod
    def generate_booking_report(cls, start_date, end_date):
        """Generate booking report for date range"""
        bookings = Booking.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        return {
            'period': f"{start_date} to {end_date}",
            'summary': {
                'total_bookings': bookings.count(),
                'total_value': float(bookings.aggregate(Sum('total_amount_usd'))['total_amount_usd__sum'] or 0),
                'avg_value': float(bookings.aggregate(Avg('total_amount_usd'))['total_amount_usd__avg'] or 0),
            },
            'by_status': list(bookings.values('status').annotate(count=Count('id'))),
            'by_aircraft': list(bookings.values('aircraft__model').annotate(
                count=Count('id'),
                revenue=Sum('total_amount_usd')
            )[:10]),
            'popular_routes': list(bookings.values(
                'departure_airport', 'arrival_airport'
            ).annotate(count=Count('id')).order_by('-count')[:10])
        }
    
    @classmethod
    def generate_customer_report(cls, start_date, end_date):
        """Generate customer analytics report"""
        users = User.objects.filter(
            date_joined__date__lte=end_date
        )
        
        new_users = users.filter(date_joined__date__gte=start_date).count()
        
        active_users = AnalyticsEvent.objects.filter(
            timestamp__date__gte=start_date,
            timestamp__date__lte=end_date
        ).values('user').distinct().count()
        
        # Customer lifetime value
        clv_data = []
        for user in users.filter(bookings__isnull=False).distinct()[:100]:
            total_spent = user.bookings.aggregate(Sum('total_amount_usd'))['total_amount_usd__sum'] or 0
            booking_count = user.bookings.count()
            clv_data.append({
                'user': user.email,
                'total_spent': float(total_spent),
                'booking_count': booking_count,
                'avg_booking': float(total_spent / booking_count) if booking_count > 0 else 0
            })
        
        return {
            'period': f"{start_date} to {end_date}",
            'summary': {
                'total_users': users.count(),
                'new_users': new_users,
                'active_users': active_users,
                'retention_rate': (active_users / users.count() * 100) if users.count() > 0 else 0,
            },
            'by_type': list(users.values('user_type').annotate(count=Count('id'))),
            'top_customers': sorted(clv_data, key=lambda x: x['total_spent'], reverse=True)[:10]
        }
    
    @classmethod
    def generate_fleet_report(cls, start_date, end_date):
        """Generate fleet utilization report"""
        from apps.fleet.models import Aircraft, FlightPerformance
        
        aircraft = Aircraft.objects.filter(is_active=True)
        
        fleet_data = []
        total_hours = 0
        
        for a in aircraft:
            flights = FlightPerformance.objects.filter(
                aircraft=a,
                flight_date__date__gte=start_date,
                flight_date__date__lte=end_date
            )
            
            hours = flights.aggregate(Sum('flight_hours'))['flight_hours__sum'] or 0
            total_hours += hours
            
            fleet_data.append({
                'registration': a.registration_number,
                'model': a.model,
                'flights': flights.count(),
                'hours': float(hours),
                'revenue': float(flights.aggregate(Sum('revenue'))['revenue__sum'] or 0),
                'utilization': (hours / (24 * 30)) * 100 if hours > 0 else 0
            })
        
        return {
            'period': f"{start_date} to {end_date}",
            'summary': {
                'total_aircraft': aircraft.count(),
                'total_flight_hours': float(total_hours),
                'avg_utilization': sum(a['utilization'] for a in fleet_data) / len(fleet_data) if fleet_data else 0,
            },
            'aircraft': sorted(fleet_data, key=lambda x: x['hours'], reverse=True)
        }
    
    @classmethod
    def generate_report(cls, report):
        """Generate report based on type"""
        if report.parameters:
            start_date = report.parameters.get('start_date')
            end_date = report.parameters.get('end_date')
        else:
            end_date = timezone.now().date()
            start_date = end_date - timezone.timedelta(days=30)
        
        if report.report_type == 'revenue':
            data = cls.generate_revenue_report(start_date, end_date)
        elif report.report_type == 'bookings':
            data = cls.generate_booking_report(start_date, end_date)
        elif report.report_type == 'customer':
            data = cls.generate_customer_report(start_date, end_date)
        elif report.report_type == 'fleet':
            data = cls.generate_fleet_report(start_date, end_date)
        else:
            data = {}
        
        return data

class ScheduledReports:
    """Handle scheduled report generation"""
    
    @classmethod
    def run_scheduled_reports(cls):
        """Run all scheduled reports due now"""
        now = timezone.now()
        reports = Report.objects.filter(
            is_active=True,
            next_run__lte=now
        )
        
        for report in reports:
            cls.run_report(report)
        
        logger.info(f"Ran {reports.count()} scheduled reports")
        return reports.count()
    
    @classmethod
    def run_report(cls, report):
        """Run a single scheduled report"""
        report.generate()
        
        # Calculate next run
        if report.frequency == 'daily':
            report.next_run = timezone.now() + timezone.timedelta(days=1)
        elif report.frequency == 'weekly':
            report.next_run = timezone.now() + timezone.timedelta(weeks=1)
        elif report.frequency == 'monthly':
            report.next_run = timezone.now() + timezone.timedelta(days=30)
        elif report.frequency == 'quarterly':
            report.next_run = timezone.now() + timezone.timedelta(days=90)
        
        report.save(update_fields=['next_run', 'last_run'])
        
        logger.info(f"Generated scheduled report {report.id}")
        return report

class ReportExporter:
    """Export reports to various formats"""
    
    @classmethod
    def to_csv(cls, data, filename):
        """Export data to CSV"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        if data and isinstance(data, list):
            writer.writerow(data[0].keys())
            
            # Write rows
            for row in data:
                writer.writerow(row.values())
        elif data and isinstance(data, dict):
            # Flatten dict for CSV
            flat_data = cls.flatten_dict(data)
            writer.writerow(flat_data.keys())
            writer.writerow(flat_data.values())
        
        return output.getvalue()
    
    @classmethod
    def to_excel(cls, data, filename):
        """Export data to Excel"""
        import openpyxl
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if data and isinstance(data, list):
            # Write headers
            for col, key in enumerate(data[0].keys(), 1):
                ws.cell(row=1, column=col, value=key)
            
            # Write data
            for row, item in enumerate(data, 2):
                for col, key in enumerate(item.keys(), 1):
                    ws.cell(row=row, column=col, value=item[key])
        elif data and isinstance(data, dict):
            flat_data = cls.flatten_dict(data)
            for col, (key, value) in enumerate(flat_data.items(), 1):
                ws.cell(row=1, column=col, value=key)
                ws.cell(row=2, column=col, value=value)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    @classmethod
    def to_json(cls, data):
        """Export data to JSON"""
        import json
        return json.dumps(data, indent=2, default=str)
    
    @classmethod
    def flatten_dict(cls, d, parent_key='', sep='_'):
        """Flatten nested dictionary"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(cls.flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)