import io
import csv
import json
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Avg
from django.utils import timezone
from django.core.files.base import ContentFile
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from .models import DailyMetric, AnalyticsEvent
from apps.bookings.models import Booking
from apps.payments.models import Payment

class ReportGenerator:
    """Generate various report types"""
    
    def __init__(self, report):
        self.report = report
        self.data = None
    
    def generate(self):
        """Generate report based on type"""
        # Gather data
        self.collect_data()
        
        # Generate file based on format
        if self.report.format == 'pdf':
            return self.generate_pdf()
        elif self.report.format == 'excel':
            return self.generate_excel()
        elif self.report.format == 'csv':
            return self.generate_csv()
        elif self.report.format == 'json':
            return self.generate_json()
        
        return None
    
    def collect_data(self):
        """Collect data for report"""
        end_date = timezone.now().date()
        
        if self.report.report_type == 'revenue':
            start_date = end_date - timedelta(days=30)
            self.data = self.get_revenue_data(start_date, end_date)
        
        elif self.report.report_type == 'bookings':
            start_date = end_date - timedelta(days=30)
            self.data = self.get_booking_data(start_date, end_date)
        
        elif self.report.report_type == 'fleet':
            self.data = self.get_fleet_data()
        
        elif self.report.report_type == 'customer':
            self.data = self.get_customer_data()
    
    def get_revenue_data(self, start_date, end_date):
        """Get revenue data"""
        payments = Payment.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            status='completed'
        )
        
        daily = DailyMetric.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date')
        
        return {
            'period': f"{start_date} to {end_date}",
            'total_revenue': payments.aggregate(Sum('amount_usd'))['amount_usd__sum'] or 0,
            'total_payments': payments.count(),
            'avg_payment': payments.aggregate(Avg('amount_usd'))['amount_usd__avg'] or 0,
            'by_method': payments.values('payment_method').annotate(
                total=Sum('amount_usd'),
                count=Count('id')
            ),
            'daily': [
                {
                    'date': m.date,
                    'revenue': m.revenue,
                    'bookings': m.total_bookings
                }
                for m in daily
            ]
        }
    
    def get_booking_data(self, start_date, end_date):
        """Get booking data"""
        bookings = Booking.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        return {
            'period': f"{start_date} to {end_date}",
            'total_bookings': bookings.count(),
            'by_status': bookings.values('status').annotate(count=Count('id')),
            'by_aircraft': bookings.values('aircraft__model').annotate(
                count=Count('id'),
                revenue=Sum('total_amount_usd')
            )[:10],
            'popular_routes': bookings.values(
                'departure_airport', 'arrival_airport'
            ).annotate(count=Count('id')).order_by('-count')[:10]
        }
    
    def get_fleet_data(self):
        """Get fleet utilization data"""
        from apps.fleet.models import Aircraft
        
        aircraft = Aircraft.objects.all()
        
        data = []
        for a in aircraft:
            bookings = a.bookings.filter(status='completed')
            total_hours = sum(
                (b.arrival_datetime - b.departure_datetime).total_seconds() / 3600
                for b in bookings
            )
            
            data.append({
                'registration': a.registration_number,
                'model': a.model,
                'total_bookings': bookings.count(),
                'flight_hours': round(total_hours, 2),
                'revenue': bookings.aggregate(Sum('total_amount_usd'))['total_amount_usd__sum'] or 0,
                'utilization_rate': min(100, (bookings.count() / 30) * 100)  # Simplified
            })
        
        return {
            'total_aircraft': len(data),
            'aircraft': sorted(data, key=lambda x: x['utilization_rate'], reverse=True)
        }
    
    def get_customer_data(self):
        """Get customer analytics data"""
        from apps.accounts.models import User
        
        users = User.objects.filter(is_active=True)
        
        return {
            'total_customers': users.count(),
            'new_this_month': users.filter(
                date_joined__month=timezone.now().month
            ).count(),
            'by_type': users.values('user_type').annotate(count=Count('id')),
            'top_customers': users.annotate(
                booking_count=Count('bookings'),
                total_spent=Sum('bookings__total_amount_usd')
            ).filter(booking_count__gt=0).order_by('-total_spent')[:10]
        }
    
    def generate_pdf(self):
        """Generate PDF report"""
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Title
        p.setFont("Helvetica-Bold", 20)
        p.drawString(50, height - 50, self.report.name)
        
        p.setFont("Helvetica", 12)
        p.drawString(50, height - 70, f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}")
        p.drawString(50, height - 85, f"Type: {self.report.get_report_type_display()}")
        
        y = height - 120
        
        # Add data based on report type
        if self.report.report_type == 'revenue' and self.data:
            y = self.add_revenue_to_pdf(p, y)
        elif self.report.report_type == 'bookings' and self.data:
            y = self.add_bookings_to_pdf(p, y)
        elif self.report.report_type == 'fleet' and self.data:
            y = self.add_fleet_to_pdf(p, y)
        
        p.save()
        buffer.seek(0)
        
        return ContentFile(
            buffer.getvalue(),
            f"{self.report.report_number}.pdf"
        )
    
    def add_revenue_to_pdf(self, p, y):
        """Add revenue data to PDF"""
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, "Revenue Summary")
        y -= 20
        
        p.setFont("Helvetica", 12)
        p.drawString(50, y, f"Period: {self.data['period']}")
        y -= 15
        p.drawString(50, y, f"Total Revenue: ${self.data['total_revenue']:,.2f}")
        y -= 15
        p.drawString(50, y, f"Total Payments: {self.data['total_payments']}")
        y -= 15
        p.drawString(50, y, f"Average Payment: ${self.data['avg_payment']:,.2f}")
        y -= 30
        
        # Daily breakdown
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "Daily Breakdown")
        y -= 20
        
        data = [['Date', 'Revenue', 'Bookings']]
        for day in self.data['daily'][:15]:  # Limit to 15 days for PDF
            data.append([
                day['date'].strftime('%Y-%m-%d'),
                f"${day['revenue']:,.2f}",
                str(day['bookings'])
            ])
        
        table = Table(data, colWidths=[100, 100, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        table.wrapOn(p, width - 100, y)
        table.drawOn(p, 50, y - (len(data) * 20))
        
        return y - (len(data) * 20) - 30
    
    def add_bookings_to_pdf(self, p, y):
        """Add booking data to PDF"""
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, "Bookings Summary")
        y -= 20
        
        p.setFont("Helvetica", 12)
        p.drawString(50, y, f"Period: {self.data['period']}")
        y -= 15
        p.drawString(50, y, f"Total Bookings: {self.data['total_bookings']}")
        y -= 30
        
        # Popular routes
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "Top Routes")
        y -= 20
        
        data = [['Route', 'Bookings']]
        for route in self.data['popular_routes'][:10]:
            data.append([
                f"{route['departure_airport']} → {route['arrival_airport']}",
                str(route['count'])
            ])
        
        table = Table(data, colWidths=[150, 80])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        table.wrapOn(p, width - 100, y)
        table.drawOn(p, 50, y - (len(data) * 20))
        
        return y - (len(data) * 20) - 30
    
    def add_fleet_to_pdf(self, p, y):
        """Add fleet data to PDF"""
        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, "Fleet Utilization")
        y -= 20
        
        p.setFont("Helvetica", 12)
        p.drawString(50, y, f"Total Aircraft: {self.data['total_aircraft']}")
        y -= 30
        
        # Aircraft breakdown
        p.setFont("Helvetica-Bold", 12)
        p.drawString(50, y, "Aircraft Details")
        y -= 20
        
        data = [['Registration', 'Model', 'Bookings', 'Hours', 'Utilization']]
        for a in self.data['aircraft'][:15]:
            data.append([
                a['registration'],
                a['model'],
                str(a['total_bookings']),
                str(a['flight_hours']),
                f"{a['utilization_rate']:.1f}%"
            ])
        
        table = Table(data, colWidths=[80, 80, 60, 60, 70])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        table.wrapOn(p, width - 100, y)
        table.drawOn(p, 50, y - (len(data) * 20))
        
        return y - (len(data) * 20) - 30
    
    def generate_excel(self):
        """Generate Excel report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.report.name[:30]
        
        # Add headers
        if self.report.report_type == 'revenue' and self.data:
            self.add_revenue_to_excel(ws)
        elif self.report.report_type == 'bookings' and self.data:
            self.add_bookings_to_excel(ws)
        elif self.report.report_type == 'fleet' and self.data:
            self.add_fleet_to_excel(ws)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return ContentFile(
            buffer.getvalue(),
            f"{self.report.report_number}.xlsx"
        )
    
    def add_revenue_to_excel(self, ws):
        """Add revenue data to Excel"""
        ws['A1'] = 'Revenue Report'
        ws['A2'] = f"Period: {self.data['period']}"
        ws['A3'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        
        ws['A5'] = 'Summary'
        ws['A6'] = 'Total Revenue'
        ws['B6'] = self.data['total_revenue']
        ws['A7'] = 'Total Payments'
        ws['B7'] = self.data['total_payments']
        ws['A8'] = 'Average Payment'
        ws['B8'] = self.data['avg_payment']
        
        ws['A10'] = 'Daily Breakdown'
        ws['A11'] = 'Date'
        ws['B11'] = 'Revenue'
        ws['C11'] = 'Bookings'
        
        row = 12
        for day in self.data['daily']:
            ws[f'A{row}'] = day['date'].strftime('%Y-%m-%d')
            ws[f'B{row}'] = day['revenue']
            ws[f'C{row}'] = day['bookings']
            row += 1
    
    def add_bookings_to_excel(self, ws):
        """Add booking data to Excel"""
        ws['A1'] = 'Bookings Report'
        ws['A2'] = f"Period: {self.data['period']}"
        ws['A3'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        
        ws['A5'] = 'Summary'
        ws['A6'] = 'Total Bookings'
        ws['B6'] = self.data['total_bookings']
        
        ws['A8'] = 'Popular Routes'
        ws['A9'] = 'Route'
        ws['B9'] = 'Bookings'
        
        row = 10
        for route in self.data['popular_routes']:
            ws[f'A{row}'] = f"{route['departure_airport']} → {route['arrival_airport']}"
            ws[f'B{row}'] = route['count']
            row += 1
    
    def add_fleet_to_excel(self, ws):
        """Add fleet data to Excel"""
        ws['A1'] = 'Fleet Utilization Report'
        ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        
        ws['A4'] = 'Aircraft Details'
        ws['A5'] = 'Registration'
        ws['B5'] = 'Model'
        ws['C5'] = 'Bookings'
        ws['D5'] = 'Flight Hours'
        ws['E5'] = 'Revenue'
        ws['F5'] = 'Utilization'
        
        row = 6
        for a in self.data['aircraft']:
            ws[f'A{row}'] = a['registration']
            ws[f'B{row}'] = a['model']
            ws[f'C{row}'] = a['total_bookings']
            ws[f'D{row}'] = a['flight_hours']
            ws[f'E{row}'] = a['revenue']
            ws[f'F{row}'] = f"{a['utilization_rate']:.1f}%"
            row += 1
    
    def generate_csv(self):
        """Generate CSV report"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        if self.report.report_type == 'revenue' and self.data:
            writer.writerow(['Date', 'Revenue', 'Bookings'])
            for day in self.data['daily']:
                writer.writerow([
                    day['date'].strftime('%Y-%m-%d'),
                    day['revenue'],
                    day['bookings']
                ])
        
        elif self.report.report_type == 'bookings' and self.data:
            writer.writerow(['Route', 'Bookings'])
            for route in self.data['popular_routes']:
                writer.writerow([
                    f"{route['departure_airport']}→{route['arrival_airport']}",
                    route['count']
                ])
        
        elif self.report.report_type == 'fleet' and self.data:
            writer.writerow(['Registration', 'Model', 'Bookings', 'Hours', 'Revenue', 'Utilization'])
            for a in self.data['aircraft']:
                writer.writerow([
                    a['registration'],
                    a['model'],
                    a['total_bookings'],
                    a['flight_hours'],
                    a['revenue'],
                    f"{a['utilization_rate']:.1f}%"
                ])
        
        content = output.getvalue()
        return ContentFile(
            content.encode('utf-8'),
            f"{self.report.report_number}.csv"
        )
    
    def generate_json(self):
        """Generate JSON report"""
        content = json.dumps(self.data, indent=2, default=str)
        return ContentFile(
            content.encode('utf-8'),
            f"{self.report.report_number}.json"
        )